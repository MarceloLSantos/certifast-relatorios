import streamlit as st

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# Criando um DataFrame de exemplo
data = {'Nome': ['João', 'Maria', 'Pedro'],
           'Idade': [25, 30, 28]}
df = pd.DataFrame(data)

# Criando um novo arquivo Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

# Escrevendo os cabeçalhos das colunas
for col_num, value in enumerate(df.columns):
    sheet.cell(row=1, column=col_num+1, value=value)

# Escrevendo os dados
for row_index, row in df.iterrows():
    for col_index, cell_value in enumerate(row):
        column_letter = get_column_letter(col_index+1)
        sheet.cell(row=row_index+2, column=col_index+1, value=cell_value)

# Salvando o arquivo
workbook.save('novo_arquivo.xlsx')